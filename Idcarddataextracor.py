# Hybrid ROI & Full-Text OCR Script with Image Normalization and Grayscale Processing
# This script uses ROI-based extraction for ALL specified fields (Name, Father Name, Gender,
# Identity Number, Dates, Country of Stay).
# It includes robust image normalization for varied input conditions (distant/close-up).
# NEW: Optimized Tesseract OCR Engine Mode (--oem 1) for potentially better recognition.

import cv2
import pytesseract
import numpy as np
import os
import re
import tempfile
from PIL import Image
from matplotlib import pyplot as plt
import openpyxl


class OCRPreprocessor:
    """
    OCR preprocessing class implementing a focused and robust pipeline.
    This version includes image normalization.
    """

    def __init__(self, show_steps=False):
        self.show_steps = show_steps

    def display_image(self, image, title="Image"):
        """Display image if show_steps is enabled"""
        if self.show_steps:
            plt.figure(figsize=(10, 6))
            if len(image.shape) == 3:
                plt.imshow(cv2.cvtColor(image, cv2.COLOR_BGR2RGB))
            else:
                plt.imshow(image, cmap='gray')
            plt.title(title)
            plt.axis('off')
            plt.show(block=False) # Use non-blocking show
            plt.pause(0.1) # Pause briefly to display the image
            plt.close('all') # Close all figures immediately after display to prevent errors

    def order_points(self, pts):
        """
        Orders a list of 4 points in top-left, top-right, bottom-right, bottom-left order.
        """
        rect = np.zeros((4, 2), dtype="float32")
        s = pts.sum(axis=1)
        rect[0] = pts[np.argmin(s)]  # Top-left has the smallest sum
        rect[2] = pts[np.argmax(s)]  # Bottom-right has the largest sum

        diff = np.diff(pts, axis=1)
        rect[1] = pts[np.argmin(diff)] # Top-right has the smallest difference
        rect[3] = pts[np.argmax(diff)] # Bottom-left has the largest difference
        return rect

    def normalize_card_image(self, image_path, target_width, target_height):
        """
        Normalizes the card image by detecting its boundaries, correcting perspective,
        and scaling it to a consistent size. Returns a grayscale normalized image.
        Handles cases where no clear 4-point contour is found (e.g., close-up images).
        Uses provided target_width and target_height for normalization.
        """
        # print("Step 0: Normalizing card image (perspective correction and scaling)...") # Removed verbose print

        # Load the image
        image = cv2.imread(image_path)
        if image is None:
            raise ValueError(f"Could not load image for normalization: {image_path}")

        # Ensure the image is 3-channel (BGR) for consistent contour detection and drawing.
        if len(image.shape) == 2:
            image_for_processing = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
            # print("   INFO: Input image was grayscale, temporarily converted to BGR for contour detection.") # Removed verbose print
        else:
            image_for_processing = image.copy()

        # Convert to grayscale for contour detection
        gray_for_contour = cv2.cvtColor(image_for_processing, cv2.COLOR_BGR2GRAY)
        
        # Apply Gaussian blur to reduce noise and help with edge detection
        blurred = cv2.GaussianBlur(gray_for_contour, (9, 9), 0) # Increased blur for smoother edges

        # Use Canny edge detection
        edged = cv2.Canny(blurred, 50, 150) # Adjusted Canny thresholds

        # Find contours in the edged image
        contours, _ = cv2.findContours(edged.copy(), cv2.RETR_LIST, cv2.CHAIN_APPROX_SIMPLE)
        
        # Sort the contours by area in descending order and keep only the largest ones
        contours = sorted(contours, key=cv2.contourArea, reverse=True)[:5]

        screenCnt = None
        # Loop over the contours
        for c in contours:
            # Approximate the contour
            peri = cv2.arcLength(c, True)
            approx = cv2.approxPolyDP(c, 0.02 * peri, True)

            # If our approximated contour has four points, then we can assume
            # that we have found our screen
            if len(approx) == 4:
                # Check if the contour area is a significant portion of the image area
                contour_area = cv2.contourArea(c)
                image_area = image.shape[0] * image.shape[1]
                # Reduced threshold for distant images: now considers contours larger than 10% of image area
                if contour_area > 0.10 * image_area: 
                    screenCnt = approx
                    break
                # else:
                    # print(f"   DEBUG: Found 4-point contour, but area ({contour_area/image_area:.2f}%) is too small. Skipping.") # Removed verbose print
        
        # --- Debugging: Display detected contour ---
        if self.show_steps:
            debug_image = image_for_processing.copy()
            if screenCnt is not None:
                cv2.drawContours(debug_image, [screenCnt], -1, (0, 255, 0), 3) # Draw green contour
                # print("   DEBUG: Detected contour drawn in green.") # Removed verbose print
            # else:
                # print("   DEBUG: No significant 4-point contour detected.") # Removed verbose print
            self.display_image(debug_image, "Original Image with Detected Contour")
        # ------------------------------------------

        normalized_image_gray = None

        if screenCnt is None:
            print("   ⚠️ Warning: Could not find a clear, significant 4-point contour for the card. Assuming entire image is the card.")
            # If no contour found, assume the entire image is the card and resize it to the provided target dimensions
            normalized_image_bgr = cv2.resize(image_for_processing, 
                                              (target_width, target_height), 
                                              interpolation=cv2.INTER_AREA)
            normalized_image_gray = cv2.cvtColor(normalized_image_bgr, cv2.COLOR_BGR2GRAY)
            # print(f"   ✅ Image resized to {target_width}x{target_height} pixels (grayscale) as fallback.") # Removed verbose print
        else:
            pts = screenCnt.reshape(4, 2)
            rect = self.order_points(pts)

            # Set up destination points for the perspective transform using provided target dimensions
            dst = np.array([
                [0, 0],
                [target_width - 1, 0],
                [target_width - 1, target_height - 1],
                [0, target_height - 1]], dtype="float32")

            M = cv2.getPerspectiveTransform(rect, dst)
            
            # Apply perspective transform to the original BGR image, then convert to grayscale
            normalized_image_bgr = cv2.warpPerspective(image_for_processing, M, (target_width, target_height))
            
            if normalized_image_bgr is None or normalized_image_bgr.size == 0:
                print("   ❌ Error: Normalized image is empty after perspective transform. Falling back to original grayscale.")
                normalized_image_gray = gray_for_contour # Fallback to original grayscale if transform fails
            else:
                normalized_image_gray = cv2.cvtColor(normalized_image_bgr, cv2.COLOR_BGR2GRAY)
            
            # print(f"   ✅ Card normalized to {target_width}x{target_height} pixels (grayscale) with perspective correction.") # Removed verbose print

        self.display_image(normalized_image_gray, "Step 0: Normalized Card Image (Grayscale)")
        return normalized_image_gray

    def step1_normalization(self, image):
        """
        Step 1: Normalization - Normalize pixel intensity values to 0-255 range.
        """
        # print("Step 1: Applying pixel intensity normalization...") # Removed verbose print
        norm_img = np.zeros((image.shape[0], image.shape[1]))
        normalized = cv2.normalize(image, norm_img, 0, 255, cv2.NORM_MINMAX)
        # self.display_image(normalized, "Step 1: Normalized Image (Pixel Intensity)") # Removed intermediate display
        return normalized

    def step2_image_scaling(self, image_path):
        """
        Placeholder for image scaling, handled by normalize_card_image.
        Returns the image as is.
        """
        # print("Step 2: Image scaling (handled by card normalization). Skipping this step.") # Removed verbose print
        return cv2.imread(image_path) 

    def step3_noise_removal(self, image):
        """
        Step 3: Noise Removal - Remove noise while preserving text.
        """
        # print("Step 3: Applying noise removal...") # Removed verbose print
        # Denoising for grayscale image
        denoised = cv2.fastNlMeansDenoising(image, None, 10, 7, 21)
        # self.display_image(denoised, "Step 3: Denoised Image") # Removed intermediate display
        return denoised

    def step4_thinning_skeletonization(self, image):
        """
        Step 4: Thinning and Skeletonization - Uniform stroke width.
        """
        # print("Step 4: Applying thinning and skeletonization...") # Removed verbose print
        # Image is already grayscale from previous steps
        gray = image.copy() 
        kernel = np.ones((2,2), np.uint8)
        opened = cv2.morphologyEx(gray, cv2.MORPH_OPEN, kernel, iterations=1)
        kernel_close = np.ones((3,3), np.uint8)
        thinned = cv2.morphologyEx(opened, cv2.MORPH_CLOSE, kernel_close, iterations=1)
        # self.display_image(thinned, "Step 4: Thinned Image") # Removed intermediate display
        return thinned

    def step5_grayscale(self, image):
        """
        Step 5: Convert to Grayscale.
        (This step will largely be a no-op if the image is already grayscale from normalization)
        """
        # print("Step 5: Converting to grayscale (if not already)...") # Removed verbose print
        if len(image.shape) == 3:
            gray = cv2.cvtColor(image, cv2.COLOR_BGR2GRAY)
        else:
            gray = image.copy() # Already grayscale
        # self.display_image(gray, "Step 5: Grayscale Image") # Removed intermediate display
        return gray

    def process_image(self, image_path, target_width, target_height):
        """
        Apply a 5-step preprocessing pipeline, returning the grayscale image
        without deskewing or a final thresholding step.
        """
        print(f"Starting preprocessing for: {image_path}")
        print("=" * 60)

        # Step 0: Normalize the card image (perspective correction and scaling), returns grayscale
        normalized_card_gray = self.normalize_card_image(image_path, target_width, target_height)

        # Now apply the rest of the preprocessing steps to the normalized grayscale card
        
        # Step 1: Normalization (pixel intensity) - applied to grayscale
        normalized = self.step1_normalization(normalized_card_gray)

        # Step 2: Noise Removal - applied to grayscale
        denoised = self.step3_noise_removal(normalized)

        # Step 3: Grayscale conversion (already grayscale, so this step will just copy)
        gray = self.step5_grayscale(denoised) 

        # Step 4: Thinning and Skeletonization - applied to grayscale
        final_image = self.step4_thinning_skeletonization(gray)

        # Display the final preprocessed image only if show_steps is True
        if self.show_steps:
            self.display_image(final_image, "Final Preprocessed Image")

        print("=" * 60)
        print("✅ All preprocessing steps completed! (Excluding deskewing, thresholding, CLAHE, and sharpening)")
        
        return final_image


class ROIExtractor:
    """
    Class to handle the extraction of text from defined Regions of Interest (ROIs).
    """
    def __init__(self, rois):
        self.rois = rois

    def extract_from_rois(self, preprocessed_image):
        """
        Iterate through ROIs and extract text from each one.
        """
        extracted_data = {}
        text_extractor = OCRTextExtractor()

        for field_name, coords in self.rois.items():
            # print(f"\n🔍 Processing ROI for: {field_name}") # Removed verbose print
            x, y, w, h = coords
            
            # Ensure coordinates are within image bounds
            img_h, img_w = preprocessed_image.shape[:2]
            x = max(0, x)
            y = max(0, y)
            w = min(w, img_w - x)
            h = min(h, img_h - y)

            if w <= 0 or h <= 0:
                print(f"   ❌ Invalid ROI dimensions for {field_name}: ({x}, {y}, {w}, {h}). Skipping.")
                extracted_data[field_name] = "Not Found"
                continue

            # Crop the image to the specified ROI
            roi_image = preprocessed_image[y:y+h, x:x+w]

            # Run OCR on the cropped ROI
            text, confidence = text_extractor.extract_text_from_image(roi_image)

            # --- Debugging: Print raw OCR text for this ROI ---
            print(f"   DEBUG: Raw OCR text for {field_name}: '{text.strip()}' (Confidence: {confidence:.1f}%)")
            # --------------------------------------------------

            # Clean the extracted text and store it
            if text:
                cleaned_text = self.clean_text(text, field_name)
                extracted_data[field_name] = cleaned_text
                print(f"   ✅ Extracted: '{cleaned_text}' (Confidence: {confidence:.1f}%)")
            else:
                extracted_data[field_name] = "Not Found"
                print(f"   ❌ No text found in ROI for: {field_name}")

        return extracted_data

    def clean_text(self, text, field_name):
        """
        Apply simple cleaning rules based on the field name.
        """
        # Basic cleanup
        text = text.replace('\n', ' ').strip()
        text = re.sub(r'^\W+|\W+$', '', text) # Remove leading/trailing non-alphanumeric chars

        # Specific cleaning for certain fields
        if field_name == "Identity Number":
            # Remove any non-digit characters from the beginning, then keep only digits and hyphens
            text = re.sub(r'^[^\d]*', '', text) # Remove leading non-digits
            text = re.sub(r'[^\d\-]', '', text)
            text = text.replace('--', '-')
        elif field_name == "Country of Stay":
            # Remove "Country of Stay", "SQuntry of Stay", or similar prefixes, case-insensitive
            # Then, ensure it starts with an alphabet, remove leading non-alpha
            text = re.sub(r'^(Country\s*of\s*Stay|SQuntry\s*of\s*Stay)\s*', '', text, flags=re.IGNORECASE).strip()
            text = re.sub(r'^[^a-zA-Z]*', '', text).strip() # Ensure it starts with an alphabet
        elif "Date" in field_name:
            text = re.sub(r'[^\d\-\./]', '', text)
        elif field_name == "Gender": # More robust gender cleaning
            # Look for 'M' or 'F' possibly followed by other characters, or 'Male'/'Female'
            # Added more specific regex to capture single 'F' or 'M' more reliably
            match = re.search(r'\b(F|M)\b|\b(Female|Male)\b', text, re.IGNORECASE)
            if match:
                if match.group(1): # Matched 'F' or 'M'
                    gender_char = match.group(1).upper()
                    if gender_char == 'M':
                        return 'Male'
                    elif gender_char == 'F':
                        return 'Female'
                elif match.group(2): # Matched 'Female' or 'Male'
                    return match.group(2).capitalize()
            return "Not Found"
        
        return text

class OCRTextExtractor:
    """
    OCR text extraction with multiple configurations
    """

    def __init__(self):
        # OCR configurations for different scenarios
        # Added --oem 1 to explicitly try the LSTM engine
        self.ocr_configs = [
            '--psm 3 --oem 1',  # Fully automatic page segmentation (general purpose) with LSTM
            '--psm 6 --oem 1',  # Assume a single uniform block of text with LSTM
            '--psm 8 --oem 1',  # Assume a single word with LSTM
            '--psm 11 --oem 1', # Sparse text. Find as much text as possible in no particular order with LSTM.
            '--psm 10 --oem 1', # Treat the image as a single character with LSTM.
            '--psm 1 --oem 1',  # Automatic page segmentation with OSD with LSTM
            '--psm 12 --oem 1', # Sparse text, with OSD with LSTM
            # Also keep original configs without explicit OEM for comparison/fallback
            '--psm 3',
            '--psm 6',
            '--psm 8',
            '--psm 11',
            '--psm 10',
            '--psm 1',
            '--psm 12'
        ]

    def extract_text_from_image(self, image, method_name=""):
        """
        Extract text using multiple OCR configurations and return the best one.
        """
        best_text = ""
        best_confidence = 0

        # Ensure image is a PIL Image object
        if isinstance(image, np.ndarray):
            pil_image = Image.fromarray(image)
        else:
            pil_image = image

        for config in self.ocr_configs:
            try:
                # Get confidence data
                data = pytesseract.image_to_data(pil_image, config=config,
                                                 output_type=pytesseract.Output.DICT)

                # Calculate average confidence
                confidences = [int(conf) for conf in data['conf'] if int(conf) > 0]
                if not confidences:
                    continue

                avg_confidence = sum(confidences) / len(confidences)
                text = pytesseract.image_to_string(pil_image, config=config).strip()

                if avg_confidence > best_confidence and len(text) > len(best_text):
                    best_confidence = avg_confidence
                    best_text = text

            except Exception as e:
                # print(f"Error with config {config}: {e}") # Optional for debugging
                continue

        return best_text, best_confidence

def save_to_excel(data, filename="idcard_data.xlsx"):
    """
    Save extracted data to Excel file
    """
    try:
        if os.path.exists(filename):
            workbook = openpyxl.load_workbook(filename)
            sheet = workbook.active
        else:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Card Data"

            # Add headers
            headers = list(data.keys())
            for col_num, header in enumerate(headers, 1):
                sheet.cell(row=1, column=col_num, value=header)

        # Add data to new row
        next_row = sheet.max_row + 1
        values = list(data.values())
        for col_num, value in enumerate(values, 1):
            sheet.cell(row=next_row, column=col_num, value=value)

        workbook.save(filename)
        print(f"✅ Data saved to: {filename}")

    except Exception as e:
        print(f"❌ Error saving to Excel: {e}")

class CardDataExtractor:
    """
    Extract structured data from OCR text, combining ROI-based and full-text pattern matching.
    """

    def extract_card_data(self, roi_extracted_data, full_image_text):
        """
        Extract structured data from raw OCR text using improved patterns.
        Combines ROI-based data with full-text pattern matching.
        """
        data = {
            "Name": "Not Found",
            "Father Name": "Not Found",
            "Gender": "Not Found",
            "Country of Stay": "Not Found",
            "Identity Number": "Not Found",
            "Date of Birth": "Not Found",
            "Date of Issue": "Not Found",
            "Date of Expiry": "Not Found"
        }

        # Populate data from ROI-extracted results for ALL fields
        for field, value in roi_extracted_data.items():
            if value != "Not Found":
                data[field] = value

        # --- Cleaning for Name and Father Name ---
        for field in ["Name", "Father Name"]:
            if data[field] != "Not Found":
                # Remove common leading labels/prefixes and then any non-alphabetic leading characters
                cleaned_name = re.sub(r'^(Name|Father\s*Name|Fault\s*Wallet|Wallet)\s*', '', data[field], flags=re.IGNORECASE).strip()
                # Ensure it starts with an alphabet, remove leading non-alpha
                cleaned_name = re.sub(r'^[^a-zA-Z]*', '', cleaned_name).strip()
                data[field] = cleaned_name
        # --- End Cleaning ---

        # No full_text_patterns or general date fallback needed as all fields are now ROI-based
        # The 'full_image_text' is still generated but not used for extraction here,
        # it's kept for debugging purposes (the Raw Full Image OCR Text print).

        return data

def main():
    """
    Main function to run the ROI-based OCR pipeline.
    """
    # ===========================================================================
    # 📌 CONFIGURATION
    # ===========================================================================
    IMAGE_PATH = r'C:\Users\Hp\Documents\orc\data\test2.jpg'  # 🛑 UPDATE THIS PATH
    SHOW_PREPROCESSING_STEPS = True  # Set to False to hide intermediate images

    # 📌 USER-DEFINED TARGET DIMENSIONS FOR NORMALIZED CARD
    # Set these to the desired width and height of your normalized card image.
    # These values will be used for both perspective-corrected and fallback images.
    USER_DEFINED_TARGET_WIDTH = 1600  # Example: 1024 pixels wide
    USER_DEFINED_TARGET_HEIGHT = 1000  # Example: 648 pixels high (for ~1.58:1 aspect ratio)

    # 📌 DEFINE YOUR REGIONS OF INTEREST (ROIs) HERE
    # Format: "Field Name": (x, y, width, height)
    # IMPORTANT: These coordinates should be based on the *normalized* card image
    # (i.e., the size defined by USER_DEFINED_TARGET_WIDTH and USER_DEFINED_TARGET_HEIGHT).
    # You will need to determine these coordinates by inspecting a normalized card image.
    # The example coordinates below are placeholders for a normalized card.
    CARD_ROIS = {
        "Name": (395, 189, 724, 210),
        "Father Name": (402, 405, 716, 192),
        "Gender": (397, 604,169,123),
        "Identity Number": (413, 739, 399, 104),  # Placeholder: Adjust these coordinates!
        "Date of Birth": (819, 728,319,116),    # Placeholder: Adjust these coordinates!
        "Date of Issue": (409, 851, 397, 113),    # Placeholder: Adjust these coordinates!
        "Date of Expiry": (817, 847, 332, 129),  # Placeholder: Adjust these coordinates!
        "Country of Stay": (574,606,545, 111)   # Placeholder: Adjust these coordinates!
    }
    # ===========================================================================

    # Check if image exists
    if not os.path.exists(IMAGE_PATH):
        print(f"❌ Image file not found: {IMAGE_PATH}")
        print("Please update the IMAGE_PATH variable in the script.")
        return

    # Continue with OCR using the defined ROIs
    print("🚀 Starting Hybrid ROI & Full-Text OCR Pipeline (now fully ROI-based for all fields)")
    print(f"📷 Input Image: {IMAGE_PATH}")
    print(f"👁️   Show Steps: {SHOW_PREPROCESSING_STEPS}")
    print("\n" + "="*80)

    try:
        # Step 1: Initialize and preprocess the full image (including normalization)
        preprocessor = OCRPreprocessor(show_steps=SHOW_PREPROCESSING_STEPS)
        preprocessed_image = preprocessor.process_image(
            IMAGE_PATH, USER_DEFINED_TARGET_WIDTH, USER_DEFINED_TARGET_HEIGHT
        )

        # Step 2: Extract text from defined ROIs for ALL fields
        roi_extractor = ROIExtractor(CARD_ROIS)
        roi_extracted_data = roi_extractor.extract_from_rois(preprocessed_image)

        # Step 3: Perform full OCR on the preprocessed image (for debug log only, not for extraction logic)
        full_text_extractor = OCRTextExtractor() 
        full_image_text, full_text_confidence = full_text_extractor.extract_text_from_image(preprocessed_image)
        print(f"\n🔍 Full Image OCR Confidence (for debug): {full_text_confidence:.1f}%")

        # --- NEW DEBUG PRINT: Raw Full Image OCR Text ---
        print(f"\n📄 Raw Full Image OCR Text (from OCRTextExtractor - for debug):")
        print("-" * 40)
        print(full_image_text)
        print("-" * 40)
        # --- END NEW DEBUG PRINT ---


        # Step 4: Extract structured data, now solely from ROI results
        data_extractor = CardDataExtractor()
        extracted_data = data_extractor.extract_card_data(roi_extracted_data, full_image_text) # full_image_text is ignored by CardDataExtractor now

        # Step 5: Display results
        print(f"\n📊 FINAL EXTRACTED CARD DATA")
        print("="*50)
        success_count = 0
        # Count total fields for success rate calculation, including those not in ROIs
        all_expected_fields = [
            "Name", "Father Name", "Gender", "Country of Stay",
            "Identity Number", "Date of Birth", "Date of Issue", "Date of Expiry"
        ]
        
        for field in all_expected_fields:
            value = extracted_data.get(field, "Not Found")
            status = "✅" if value != "Not Found" and value else "❌"
            if value and value != "Not Found":
                success_count += 1
            print(f"{status} {field:15}: {value}")

        print("="*50)
        print(f"📈 Success Rate: {success_count}/{len(all_expected_fields)} fields ({success_count/len(all_expected_fields)*100:.1f}%)")

        # Step 6: Save to Excel
        save_to_excel(extracted_data)

        # Step 7: Save debug information
       # with open("ocr_debug_log_hybrid.txt", "w", encoding="utf-8") as f:
           # f.write("Hybrid ROI & Full-Text OCR Debug Log\n")
            #f.write("====================================\n")
            #f.write(f"Full Image OCR Confidence: {full_text_confidence:.1f}%\n")
            #f.write(f"Success Rate: {success_count}/{len(all_expected_fields)} ({success_count/len(all_expected_fields)*100:.1f}%)\n\n")
            #f.write("Extracted Data:\n")
            #f.write("-" * 40 + "\n")
            #for field, value in extracted_data.items():
            #    f.write(f"{field}: {value}\n")
            #f.write("\nRaw Full Image OCR Text:\n")
            #f.write("-" * 40 + "\n")
            #f.write(full_image_text)
            #f.write("\n" + "-" * 40 + "\n")

        #print("💾 Debug log saved to: ocr_debug_log_hybrid.txt")

    except Exception as e:
        print(f"\n❌ Error during processing: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
